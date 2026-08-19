"""
Estudio Analítico de Pujas y Mercado (Test 99 - Pujas)
======================================================
Analiza el historial de subastas y transferencias de la liga a partir de:
- ./data/board_bids.csv
- ./data/board_transfers.csv
- ./data/players.csv / ./data/_master.csv

Genera:
1. Excel estructurado: ./test/99_pujas/estudio_pujas.xlsx (con 3 pestañas formateadas)
2. Informe Markdown: ./test/99_pujas/resumen_estudio_pujas.md
3. Resumen en consola con métricas clave.

Usage:
  .venv/bin/python test/99_pujas/estudio_pujas_suite.py
"""

import sys
import os
import time
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Add script directory and project root to sys.path
script_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.abspath(os.path.join(script_dir, '../..'))

if script_dir not in sys.path:
    sys.path.insert(0, script_dir)
if project_root not in sys.path:
    sys.path.insert(0, project_root)


def load_data():
    """Loads bids, transfers, and player datasets."""
    def _find(f):
        raw = os.path.join('./data/raw', f)
        return raw if os.path.exists(raw) else os.path.join('./data', f)

    bids_path = _find('board_bids.csv')
    transfers_path = _find('board_transfers.csv')
    players_path = './data/players_transformed.csv' if os.path.exists('./data/players_transformed.csv') else ('./data/_master.csv' if os.path.exists('./data/_master.csv') else _find('players.csv'))

    if not os.path.exists(bids_path) or not os.path.exists(transfers_path):
        raise FileNotFoundError("Could not find board_bids.csv or board_transfers.csv in ./data/raw/ or ./data/")

    df_bids = pd.read_csv(bids_path)
    df_transfers = pd.read_csv(transfers_path)
    df_players = pd.read_csv(players_path)

    return df_bids, df_transfers, df_players


def build_auction_study_data(df_bids, df_transfers, df_players):
    """Processes raw bid and transfer data to compute auction metrics using historical price reconstruction."""
    now_dt = datetime.datetime.now()

    # Build map for player metadata
    player_map = {}
    for _, r in df_players.iterrows():
        p_id = r.get('id') or r.get('PLAYER_ID')
        p_name = r.get('name') or r.get('PLAYER_NAME')
        p_pos = r.get('position') or r.get('PLAYER_POSITION')
        p_team = r.get('team_name') or r.get('TEAM_NAME')
        p_price = r.get('price') or r.get('PLAYER_PRICE')
        p_inc = r.get('price_increment') or r.get('PLAYER_PRICE_INCREMENT') or 0.0
        player_map[p_id] = {
            'name': str(p_name) if pd.notna(p_name) else f'ID {p_id}',
            'pos': str(p_pos) if pd.notna(p_pos) else 'UNK',
            'team': str(p_team) if pd.notna(p_team) else 'UNK',
            'price': float(p_price) if pd.notna(p_price) and float(p_price) > 0 else 0.0,
            'inc': float(p_inc) if pd.notna(p_inc) else 0.0
        }

    # Filter market buys from computer (Mercado)
    market_buys = df_transfers[df_transfers['type'] == 'market_buy'].copy()

    # Load 100% exact official daily market prices from Biwenger API if available
    hist_prices_file = os.path.join(os.path.dirname(__file__), '../../data/historical_player_prices.json')
    hist_prices_db = {}
    if os.path.exists(hist_prices_file):
        try:
            with open(hist_prices_file, 'r', encoding='utf-8') as f:
                hist_prices_db = json.load(f)
        except Exception:
            hist_prices_db = {}

    auction_records = []
    for _, row in market_buys.iterrows():
        p_id = row['player_id']
        date_str = str(row['date'])
        winner_name = str(row['buyer_name'])
        win_amount = float(row['amount'])

        p_info = player_map.get(p_id, {})
        player_name = p_info.get('name', f'ID {p_id}')
        pos = p_info.get('pos', 'UNK')
        team = p_info.get('team', 'UNK')
        curr_price = p_info.get('price', win_amount)
        daily_inc = p_info.get('inc', 0.0)

        # Match losing bids for this player around the auction date
        losing_bids = df_bids[(df_bids['player_id'] == p_id) & (df_bids['date'] == date_str)]
        num_losing_bids = len(losing_bids)
        total_bidders = num_losing_bids + 1  # Winner + losing bidders

        # Exact market price lookup from official Biwenger prices history
        # Bids are placed on Bidding Day (D-1) at Price(D-1), observing visible Daily Increment (Price(D-1) - Price(D-2))
        auc_dt_res = pd.to_datetime(date_str)
        auc_dt_bid = auc_dt_res - pd.Timedelta(days=1)
        auc_dt_prev = auc_dt_bid - pd.Timedelta(days=1)

        yymmdd_res = auc_dt_res.strftime('%y%m%d')
        yymmdd_bid = auc_dt_bid.strftime('%y%m%d')
        yymmdd_prev = auc_dt_prev.strftime('%y%m%d')

        p_prices = hist_prices_db.get(str(p_id), {})
        price_bid_day = p_prices.get(yymmdd_bid)
        price_prev_day = p_prices.get(yymmdd_prev)
        price_res_day = p_prices.get(yymmdd_res)

        if price_bid_day and float(price_bid_day) > 0:
            hist_price = float(price_bid_day)
            if price_prev_day and float(price_prev_day) > 0:
                intraday_inc_eur = float(price_bid_day) - float(price_prev_day)
                intraday_inc_pct = (intraday_inc_eur / hist_price) * 100.0
            else:
                intraday_inc_eur = daily_inc
                intraday_inc_pct = (daily_inc / hist_price) * 100.0 if hist_price > 0 else 0.0
        elif price_res_day and float(price_res_day) > 0:
            hist_price = float(price_res_day)
            intraday_inc_eur = daily_inc
            intraday_inc_pct = (daily_inc / hist_price) * 100.0 if hist_price > 0 else 0.0
        else:
            days_ago = max(0, (now_dt.date() - auc_dt_res.date()).days)
            hist_price = curr_price - (days_ago * daily_inc)
            if hist_price <= 0:
                hist_price = curr_price
            intraday_inc_eur = daily_inc
            intraday_inc_pct = (daily_inc / hist_price) * 100.0 if hist_price > 0 else 0.0

            # Min bid capping fallback: starting price cannot exceed any valid bid placed on the player
            if num_losing_bids > 0:
                min_bid_found = float(losing_bids['bid_amount'].min())
                if hist_price > min_bid_found:
                    hist_price = min_bid_found

        second_highest_bid = float(losing_bids['bid_amount'].max()) if num_losing_bids > 0 else None

        overbid_market_eur = win_amount - hist_price
        overbid_market_pct = (overbid_market_eur / hist_price * 100.0) if hist_price > 0 else 0.0

        excess_vs_2nd_eur = (win_amount - second_highest_bid) if second_highest_bid is not None else 0.0

        losing_bidders_list = [
            f"{r['bidder_name']} ({float(r['bid_amount'])/1e6:.2f}M€)"
            for _, r in losing_bids.iterrows()
        ]
        losing_bidders_str = ", ".join(losing_bidders_list) if losing_bidders_list else "Sin pujas rivales"

        auction_records.append({
            'Fecha': date_str,
            'PLAYER_ID': p_id,
            'Jugador': player_name,
            'Posición': pos,
            'Equipo': team,
            'Mánager Ganador': winner_name,
            'Precio Mercado Salida (€)': hist_price,
            'Subida Intradía (€)': intraday_inc_eur,
            'Subida Intradía (%)': intraday_inc_pct,
            'Precio Mercado Actual (€)': curr_price,
            'Precio Ganador (€)': win_amount,
            'Nº Pujadores': total_bidders,
            'Sobrepuja s/Salida (€)': overbid_market_eur,
            'Sobrepuja s/Salida (%)': overbid_market_pct,
            '2ª Puja Más Alta (€)': second_highest_bid if second_highest_bid is not None else 0.0,
            'Dinero Excesivo s/2ª Puja (€)': excess_vs_2nd_eur,
            'Pujadores Rivales': losing_bidders_str
        })

    df_detail = pd.DataFrame(auction_records)

    # --------------------------------------------------------------------------
    # MANAGER SUMMARY AGGREGATION
    # --------------------------------------------------------------------------
    manager_summary = []
    for manager, group in df_detail.groupby('Mánager Ganador'):
        total_won = len(group)
        total_spent = group['Precio Ganador (€)'].sum()
        avg_price = group['Precio Ganador (€)'].mean()
        avg_overbid_pct = group['Sobrepuja s/Salida (%)'].mean()
        total_excess_vs_2nd = group['Dinero Excesivo s/2ª Puja (€)'].sum()

        manager_summary.append({
            'Mánager': manager,
            'Subastas Ganadas': total_won,
            'Inversión Total (€)': total_spent,
            'Precio Medio (€)': avg_price,
            'Sobrepuja Media s/Salida (%)': avg_overbid_pct,
            'Exceso de Puja s/2ª Opción (€)': total_excess_vs_2nd
        })

    df_managers = pd.DataFrame(manager_summary).sort_values(by='Subastas Ganadas', ascending=False)

    # --------------------------------------------------------------------------
    # COMPETITION TIER MATRIX
    # --------------------------------------------------------------------------
    tier_summary = []
    tier_groups = {
        '1 Pujador (Sin competencia)': df_detail[df_detail['Nº Pujadores'] == 1],
        '2 Pujadores (Competencia moderada)': df_detail[df_detail['Nº Pujadores'] == 2],
        '3+ Pujadores (Guerra de subasta)': df_detail[df_detail['Nº Pujadores'] >= 3],
    }

    for tier_label, sub_df in tier_groups.items():
        count = len(sub_df)
        avg_overbid_pct = sub_df['Sobrepuja s/Salida (%)'].mean() if count > 0 else 0.0
        avg_overbid_eur = sub_df['Sobrepuja s/Salida (€)'].mean() if count > 0 else 0.0
        avg_spent = sub_df['Precio Ganador (€)'].mean() if count > 0 else 0.0

        tier_summary.append({
            'Nivel de Competencia': tier_label,
            'Nº Subastas': count,
            'Precio Medio Fichaje (€)': avg_spent,
            'Sobrepuja Media (€)': avg_overbid_eur,
            'Sobrepuja Media (%)': avg_overbid_pct
        })

    df_tiers = pd.DataFrame(tier_summary)

    return df_detail, df_managers, df_tiers


def generate_formatted_excel(df_detail, df_managers, df_tiers, excel_path, df_matrix=None, df_sim_detail=None):
    """Writes results to a beautifully formatted Excel file using openpyxl."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Colors and Styles
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", size=10, bold=True)

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # 1. Sheet: Detalle_Subastas
    ws_detail = wb.create_sheet(title="Detalle_Subastas")
    ws_detail.views.sheetView[0].showGridLines = True

    headers_detail = list(df_detail.columns)
    ws_detail.append(headers_detail)

    for r_idx, row in enumerate(df_detail.itertuples(index=False), start=2):
        ws_detail.append(list(row))

    # Format Detalle_Subastas
    for col_idx in range(1, len(headers_detail) + 1):
        cell = ws_detail.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(2, len(df_detail) + 2):
        for col_idx, col_name in enumerate(headers_detail, start=1):
            cell = ws_detail.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border

            # Number formatting
            if ' (€)' in col_name:
                cell.number_format = '#,##0 €'
                cell.alignment = Alignment(horizontal="right")
            elif ' (%)' in col_name:
                cell.number_format = '0.00"%"'
                cell.alignment = Alignment(horizontal="right")
            elif col_name in ['Nº Pujadores', 'PLAYER_ID']:
                cell.alignment = Alignment(horizontal="center")

    # 2. Sheet: Resumen_Por_Rivales
    ws_managers = wb.create_sheet(title="Resumen_Por_Rivales")
    ws_managers.views.sheetView[0].showGridLines = True

    headers_managers = list(df_managers.columns)
    ws_managers.append(headers_managers)

    for row in df_managers.itertuples(index=False):
        ws_managers.append(list(row))

    for col_idx in range(1, len(headers_managers) + 1):
        cell = ws_managers.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(2, len(df_managers) + 2):
        for col_idx, col_name in enumerate(headers_managers, start=1):
            cell = ws_managers.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border

            if ' (€)' in col_name:
                cell.number_format = '#,##0 €'
                cell.alignment = Alignment(horizontal="right")
            elif ' (%)' in col_name:
                cell.number_format = '0.00"%"'
                cell.alignment = Alignment(horizontal="right")
            elif col_name == 'Subastas Ganadas':
                cell.alignment = Alignment(horizontal="center")

    # 3. Sheet: Competencia_vs_Sobrepuja
    ws_tiers = wb.create_sheet(title="Competencia_vs_Sobrepuja")
    ws_tiers.views.sheetView[0].showGridLines = True

    headers_tiers = list(df_tiers.columns)
    ws_tiers.append(headers_tiers)

    for row in df_tiers.itertuples(index=False):
        ws_tiers.append(list(row))

    for col_idx in range(1, len(headers_tiers) + 1):
        cell = ws_tiers.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(2, len(df_tiers) + 2):
        for col_idx, col_name in enumerate(headers_tiers, start=1):
            cell = ws_tiers.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border

            if ' (€)' in col_name:
                cell.number_format = '#,##0 €'
                cell.alignment = Alignment(horizontal="right")
            elif ' (%)' in col_name:
                cell.number_format = '0.00"%"'
                cell.alignment = Alignment(horizontal="right")
            elif col_name == 'Nº Subastas':
                cell.alignment = Alignment(horizontal="center")

    # 4. Sheet: Modelo_Predictivo (if provided)
    if df_matrix is not None and not df_matrix.empty:
        ws_matrix = wb.create_sheet(title="Modelo_Predictivo")
        ws_matrix.views.sheetView[0].showGridLines = True

        headers_matrix = list(df_matrix.columns)
        ws_matrix.append(headers_matrix)

        for row in df_matrix.itertuples(index=False):
            ws_matrix.append(list(row))

        for col_idx in range(1, len(headers_matrix) + 1):
            cell = ws_matrix.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx in range(2, len(df_matrix) + 2):
            for col_idx, col_name in enumerate(headers_matrix, start=1):
                cell = ws_matrix.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border

                if ' (€)' in col_name:
                    cell.number_format = '#,##0 €'
                    cell.alignment = Alignment(horizontal="right")
                elif ' (%)' in col_name:
                    cell.number_format = '0.0"%"'
                    cell.alignment = Alignment(horizontal="right")
    # 5. Sheet: Simulacion_Historica (if provided)
    if df_sim_detail is not None and not df_sim_detail.empty:
        ws_sim = wb.create_sheet(title="Simulacion_Historica")
        ws_sim.views.sheetView[0].showGridLines = True

        headers_sim = list(df_sim_detail.columns)
        ws_sim.append(headers_sim)

        for row in df_sim_detail.itertuples(index=False):
            ws_sim.append(list(row))

        for col_idx in range(1, len(headers_sim) + 1):
            cell = ws_sim.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx in range(2, len(df_sim_detail) + 2):
            for col_idx, col_name in enumerate(headers_sim, start=1):
                cell = ws_sim.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border

                if ' (€)' in col_name:
                    cell.number_format = '#,##0 €'
                    cell.alignment = Alignment(horizontal="right")
                elif ' (%)' in col_name:
                    cell.number_format = '0.00"%"'
                    cell.alignment = Alignment(horizontal="right")
                elif col_name in ['Posición', 'Resultado Modelo']:
                    cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths across all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                max_len = max(max_len, len(val_str))
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(excel_path)
    print(f"📊 Formatted Excel created successfully at:\n   📄 {excel_path}")


def generate_markdown_report(df_detail, df_managers, df_tiers, md_path):
    """Generates a rich analytical Markdown report summarizing key insights."""
    total_auctions = len(df_detail)
    total_spent = df_detail['Precio Ganador (€)'].sum()
    avg_overbid_pct = df_detail['Sobrepuja s/Salida (%)'].mean()
    total_excess_vs_2nd = df_detail['Dinero Excesivo s/2ª Puja (€)'].sum()

    top_buyer = df_managers.iloc[0] if not df_managers.empty else None
    top_overbidder = df_managers.sort_values(by='Sobrepuja Media s/Salida (%)', ascending=False).iloc[0] if not df_managers.empty else None

    md_output = []
    md_output.append("# 📈 ESTUDIO ANALÍTICO DE SUBASTAS Y SOBREPUJAS DE LA LIGA")
    md_output.append(f"**Fecha de Análisis:** {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  ")
    md_output.append(f"**Total de Subastas Analizadas:** `{total_auctions}`  ")
    md_output.append(f"**Inversión Total Acumulada:** `{total_spent/1_000_000.0:.2f}M €` ({total_spent:,.0f} €)  ")
    md_output.append(f"**Sobrepuja Media Global sobre Valor de Salida:** `{avg_overbid_pct:.2f}%`  ")
    md_output.append(f"**Exceso de Dinero Pagado sobre 2ª Oferta:** `{total_excess_vs_2nd/1_000_000.0:.2f}M €`  ")
    md_output.append("\n---\n")

    # 1. Hallazgos Clave
    md_output.append("## 🔍 1. Hallazgos Clave de Inteligencia de Mercado")
    if top_buyer is not None:
        md_output.append(f"* 👑 **Mánager Más Activo:** `{top_buyer['Mánager']}` con `{top_buyer['Subastas Ganadas']}` jugadores fichados e inversión de `{top_buyer['Inversión Total (€)']/1e6:.2f}M €`.")
    if top_overbidder is not None:
        md_output.append(f"* 🔥 **Mánager Más Agresivo (Sobrepuja %):** `{top_overbidder['Mánager']}` con una sobrepuja media del `{top_overbidder['Sobrepuja Media s/Salida (%)']:.2f}%` sobre valor de mercado en subasta.")

    md_output.append("\n---\n")

    # 2. Matriz de Competencia
    md_output.append("## ⚔️ 2. Comportamiento de Sobrepuja según Grado de Competencia")
    df_tiers_fmt = df_tiers.copy()
    df_tiers_fmt['Precio Medio Fichaje (€)'] = df_tiers_fmt['Precio Medio Fichaje (€)'].apply(lambda x: f"{x/1e6:.2f}M €")
    df_tiers_fmt['Sobrepuja Media (€)'] = df_tiers_fmt['Sobrepuja Media (€)'].apply(lambda x: f"{x/1e6:.2f}M €")
    df_tiers_fmt['Sobrepuja Media (%)'] = df_tiers_fmt['Sobrepuja Media (%)'].apply(lambda x: f"{x:.2f}%")
    md_output.append(df_tiers_fmt.to_markdown(index=False))

    md_output.append("\n---\n")

    # 3. Resumen por Rivales
    md_output.append("## 👥 3. Perfil y Comportamiento de Puja por Mánager")
    df_managers_fmt = df_managers.copy()
    df_managers_fmt['Inversión Total (€)'] = df_managers_fmt['Inversión Total (€)'].apply(lambda x: f"{x/1e6:.2f}M €")
    df_managers_fmt['Precio Medio (€)'] = df_managers_fmt['Precio Medio (€)'].apply(lambda x: f"{x/1e6:.2f}M €")
    df_managers_fmt['Sobrepuja Media s/Salida (%)'] = df_managers_fmt['Sobrepuja Media s/Salida (%)'].apply(lambda x: f"{x:.2f}%")
    df_managers_fmt['Exceso de Puja s/2ª Opción (€)'] = df_managers_fmt['Exceso de Puja s/2ª Opción (€)'].apply(lambda x: f"{x/1e6:.2f}M €")

    md_output.append(df_managers_fmt.to_markdown(index=False))

    md_output.append("\n---\n")

    # 4. Muestra de Subastas Relevantes
    md_output.append("## 📋 4. Muestra de Subastas Analizadas (Últimas 15 operadas)")
    df_sample = df_detail.head(15).copy()
    df_sample['Precio Mercado Salida (€)'] = df_sample['Precio Mercado Salida (€)'].apply(lambda x: f"{x/1e6:.2f}M €")
    df_sample['Precio Mercado Actual (€)'] = df_sample['Precio Mercado Actual (€)'].apply(lambda x: f"{x/1e6:.2f}M €")
    df_sample['Precio Ganador (€)'] = df_sample['Precio Ganador (€)'].apply(lambda x: f"{x/1e6:.2f}M €")
    df_sample['Sobrepuja s/Salida (%)'] = df_sample['Sobrepuja s/Salida (%)'].apply(lambda x: f"{x:.2f}%")
    cols_display = ['Fecha', 'Jugador', 'Posición', 'Mánager Ganador', 'Precio Mercado Salida (€)', 'Precio Mercado Actual (€)', 'Precio Ganador (€)', 'Nº Pujadores', 'Sobrepuja s/Salida (%)']
    md_output.append(df_sample[cols_display].to_markdown(index=False))

    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(md_output))

    print(f"✅ Markdown Report created successfully at:\n   📄 {md_path}")


def run_pujas_suite():
    test_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(test_dir, "estudio_pujas.xlsx")
    md_path = os.path.join(test_dir, "resumen_estudio_pujas.md")

    print("=" * 70)
    print("📊 TEST 99 — ESTUDIO ANALÍTICO DE PUJAS Y SOBREPUJAS DE LA LIGA")
    print("=" * 70)

    start_time = time.time()

    # 1. Load Data
    df_bids, df_transfers, df_players = load_data()
    print(f"✅ Data loaded: {len(df_bids)} bid logs | {len(df_transfers)} transfer logs | {len(df_players)} players")

    # 2. Build Analytics & Predictive Model
    df_detail, df_managers, df_tiers = build_auction_study_data(df_bids, df_transfers, df_players)

    from modelo_predictivo_pujas import run_predictive_model_suite
    model, df_matrix = run_predictive_model_suite()

    from simulacion_pujas import run_auction_simulation
    df_sim_detail, df_scen = run_auction_simulation()

    # 3. Export Formatted Excel (including Modelo_Predictivo & Simulacion_Historica sheets)
    generate_formatted_excel(df_detail, df_managers, df_tiers, excel_path, df_matrix=df_matrix, df_sim_detail=df_sim_detail)

    # 4. Export Markdown Report
    generate_markdown_report(df_detail, df_managers, df_tiers, md_path)

    elapsed = time.time() - start_time

    # 5. Terminal Dashboard Summary
    print("\n" + "=" * 70)
    print("📊 RESUMEN EJECUTIVO DEL ESTUDIO DE PUJAS")
    print("=" * 70)
    print(f" ⚽ Subastas Analizadas      : {len(df_detail)}")
    print(f" 💰 Inversión Total Liga     : {df_detail['Precio Ganador (€)'].sum()/1e6:.2f}M €")
    print(f" 📈 Sobrepuja Media Global  : {df_detail['Sobrepuja s/Salida (%)'].mean():.2f}%")
    print(f" ⏱️ Tiempo de Procesamiento : {elapsed:.2f} segundos")
    print(f" 🟢 Estado del Test          : SUCCESS ✅")
    print("-" * 70)
    print("📄 ARCHIVOS GENERADOS")
    print("-" * 70)
    print(f" 📊 Excel Completo (XLSX)    : {excel_path}")
    print(f" 📋 Informe Ejecutivo (MD)   : {md_path}")
    print("=" * 70 + "\n")


if __name__ == "__main__":
    run_pujas_suite()
