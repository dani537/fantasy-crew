import pandas as pd
from dotenv import load_dotenv
import os
import io
from contextlib import redirect_stdout

# Imports for data extraction
from src.tools.data_extraction.auth import BiwengerAuth
from src.tools.data_extraction.biwenger_data import BiwengerGeneralData, UserLeagueData
from src.tools.data_extraction.external_data import ComuniateData, JornadaPerfectaData, EuroClubIndexData

# Imports for data transformation
from src.tools.data_extraction.transformers import (
    print_step,
    rename_and_normalize_columns,
    process_comuniate,
    process_odds,
    consolidate_player_data,
    feature_engineering
)

# Config
from src.config import Credentials

def extract_and_save_data():
    """
    Simulates the data extraction process.
    Authenticates with Biwenger and fetches all required data, saving it to CSVs.
    """
    load_dotenv()

    missing = Credentials.validate()
    if missing:
        raise RuntimeError(f"Missing required credentials in .env: {', '.join(missing)}")

    # Suppress output from imported modules to reduce noise
    f = io.StringIO()
    
    try:
        # Step 1: General LaLiga Data (100% Public & Anonymous - No Auth)
        print_step(1, "Extracting Public General Data for Competition (Anonymous)")
        with redirect_stdout(f):
            general_data = BiwengerGeneralData(session=None, competition_slug="la-liga")
            general_data.run()
            season_info = general_data.season_info()

        # Step 2: External Data (Comuniate, Jornada Perfecta, EuroClubIndex - Anonymous)
        print_step(2, "Extracting External Public Data (Comuniate, Jornada Perfecta, EuroClubIndex)")
        with redirect_stdout(f):
            comuniate = ComuniateData()
            df_comuniate = comuniate.run()
            
            jp = JornadaPerfectaData()
            df_news = jp.run()

            eci = EuroClubIndexData()
            df_odds = eci.run()

        # Step 3: User Authentication (Only required for Private League Data)
        print_step(3, "Authenticating with Biwenger for Private League Data")
        with redirect_stdout(f):
            auth = BiwengerAuth(
                email=Credentials.BIWENGER_USERNAME,
                password=Credentials.BIWENGER_PASSWORD,
                token=Credentials.BIWENGER_TOKEN
            )
            auth.run()
        
        # Save User Info/Metadata
        if auth.player_info:
            user_info_dict = {
                'user_id': [auth.player_info.user_id],
                'user_name': [auth.player_info.user_name],
                'league_id': [auth.player_info.league_id],
                'league_name': [auth.player_info.league_name],
                'team_id': [auth.player_info.team_id],
                'team_name': [auth.player_info.team_name],
                'balance': [auth.player_info.balance]
            }
            df_user_info = pd.DataFrame(user_info_dict)
            os.makedirs('./data', exist_ok=True)
            df_user_info.to_csv('./data/user_info.csv', index=False)

        # Step 4: Private Fantasy League Data
        print_step(4, "Extracting User Private League Data (Table, Market, My Players)")
        with redirect_stdout(f):
            user_league_data = UserLeagueData(session=auth.session, token=auth.token, league_id=auth.player_info.league_id, user_id=auth.player_info.team_id)
            user_league_data.run(auth.session)

        # Step 5: Saving raw extracted data to ./data/raw/
        print_step(5, "Saving raw staging data to ./data/raw/")
        os.makedirs('./data/raw', exist_ok=True)
        os.makedirs('./data', exist_ok=True)
        
        general_data.df_players.to_csv('./data/raw/players.csv', index=False)
        general_data.df_teams.to_csv('./data/raw/teams.csv', index=False)
        general_data.df_next_jornada.to_csv('./data/raw/next_jornada.csv', index=False)
        
        pd.DataFrame(season_info.rounds).to_csv('./data/raw/rounds.csv', index=False)
        
        # Enriched active events schedule in root ./data/
        if hasattr(general_data, 'df_active_events') and not general_data.df_active_events.empty:
            general_data.df_active_events.to_csv('./data/active_events.csv', index=False)

        user_league_data.df_league_players.to_csv('./data/raw/league_players.csv', index=False)
        user_league_data.df_league_table.to_csv('./data/raw/league_teams.csv', index=False)
        user_league_data.df_market_offers.to_csv('./data/raw/market_offers.csv', index=False)
        user_league_data.df_market_sales.to_csv('./data/raw/market_sales.csv', index=False)

        if hasattr(user_league_data, 'df_board_transfers'):
            user_league_data.df_board_transfers.to_csv('./data/raw/board_transfers.csv', index=False)
        if hasattr(user_league_data, 'df_board_bids'):
            user_league_data.df_board_bids.to_csv('./data/raw/board_bids.csv', index=False)

        df_comuniate.to_csv('./data/raw/comuniate.csv', index=False)
        df_odds.to_csv('./data/raw/odds.csv', index=False)
        
        # Contextual news in ./data/
        df_news.to_csv('./data/news.csv', index=False)

        # Step 6: Market Tracker & Rivals Financials (Google Sheets + CSV)
        print_step(6, "Syncing Market Tracker (Google Sheets) & Estimating Rivals Balances")
        try:
            from src.tracker import BiwengerSheetsTracker
            tracker = BiwengerSheetsTracker()
            tracker.sync(days_back=7)
        except Exception as e:
            print(f"⚠️ Warning: Google Sheets Tracker sync skipped: {e}")
        
        print("✅ Data extraction pipeline completed successfully.")


        
    except Exception as e:
        print(f"❌ Error during extraction: {e}")
        print("--- Detailed Logs ---")
        print(f.getvalue())
        raise e

def import_data():
    """
    Reads the CSV files generated by extraction.
    Checks ./data/raw first for raw tables, falling back to ./data/.
    """
    print_step(8, "Importing data from CSVs")
    file_mapping = {
        'players': 'players.csv',
        'teams': 'teams.csv',
        'next_match': 'next_jornada.csv',
        'league_players': 'league_players.csv',
        'league_teams': 'league_teams.csv',
        'market_offers': 'market_offers.csv',
        'market_sales': 'market_sales.csv',
        'board_transfers': 'board_transfers.csv',
        'board_bids': 'board_bids.csv',
        'rival_financials': 'rival_financials.csv',
        'comuniate': 'comuniate.csv',
        'news': 'news.csv',
        'odds': 'odds.csv',
        'user_info': 'user_info.csv',
        'rounds': 'rounds.csv',
        'active_events': 'active_events.csv'
    }
    
    imported_data = {}
    for name, filename in file_mapping.items():
        root_path = os.path.join('./data', filename)
        raw_path = os.path.join('./data/raw', filename)
        path = root_path if os.path.exists(root_path) else raw_path

        if os.path.exists(path):
            try:
                if os.path.getsize(path) < 5: 
                     imported_data[name] = pd.DataFrame()
                else:
                    imported_data[name] = pd.read_csv(path)
            except pd.errors.EmptyDataError:
                imported_data[name] = pd.DataFrame() 
            except Exception as e:
                imported_data[name] = pd.DataFrame()
        else:
            imported_data[name] = pd.DataFrame()
            
    return imported_data

def format_master_excel(excel_path: str):
    """
    Applies professional styling, number formats, auto-filters, and frozen panes to _master.xlsx.
    """
    if not os.path.exists(excel_path):
        return

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.load_workbook(excel_path)

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")  # Dark Slate
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        thin_border_side = Side(border_style="thin", color="E2E8F0")
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        header_border = Border(
            left=Side(border_style="thin", color="0F172A"),
            right=Side(border_style="thin", color="0F172A"),
            top=Side(border_style="medium", color="0F172A"),
            bottom=Side(border_style="medium", color="0F172A")
        )

        data_font = Font(name="Calibri", size=10, color="1E293B")

        tab_colors = {
            "user_info": "2563EB",            # Blue
            "active_events": "0D9488",        # Teal
            "players_transformed": "059669",  # Emerald Green
            "rival_financials": "D97706",     # Amber
            "news": "7C3AED"                  # Purple
        }

        currency_cols = {
            'PLAYER_PRICE', 'MARKET_SALE_PRICE', 'BIWPLAYER_CLAUSE', 'BIWPLAYER_PURCHASE_PRICE',
            'BIWPLAYER_INVESTED', 'MARKET_OFFER_AMOUNT', 'saldo_disponible', 'valor_plantilla',
            'patrimonio_total', 'beneficio_neto', 'balance', 'total_spent', 'total_income', 'price',
            'amount', 'clause', 'teamValue'
        }
        
        diff_currency_cols = {
            'PLAYER_PRICE_INCREMENT', 'teamValueInc', 'net_balance_change'
        }

        pct_cols = {
            'COMUNIATE_STARTER', 'COMUNIATE_SUPPLENT', 'NEXT_GAME_WIN', 'PERCENTILE', 'POSITION_PERCENTILE'
        }

        decimal_cols = {
            'EXPECTED_POINTS', 'AVG_POINTS', 'AVG_POINTS_HOME', 'AVG_POINTS_AWAY',
            'AVG_POINTS_MOMENTUM', 'COST_PER_XP', 'MOMENTUM_TREND', '1', 'X', '2'
        }

        integer_cols = {
            'PLAYER_POINTS', 'PLAYER_POINTS_HOME', 'PLAYER_POINTS_AWAY', 'puntos',
            'position', 'teamSize', 'num_jugadores', 'bids_placed', 'purchases',
            'sales', 'clauses_paid', 'clauses_received', 'TOTAL_PARTIDOS', 'ID_JORNADA'
        }

        center_cols = {
            'PLAYER_ID', 'TEAM_ID', 'PLAYER_POSITION', 'PLAYER_STATUS', 'NEXT_GAME',
            'IS_MY_PLAYER', 'ON_MARKET', 'posicion', 'en_deuda', 'amenaza_clausulazo',
            'id', 'user_id', 'league_id', 'team_id', 'ID_JORNADA', 'ESTADO', 'FECHA_INICIO', 'FECHA_FIN'
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            if sheet_name in tab_colors:
                ws.sheet_properties.tabColor = tab_colors[sheet_name]

            ws.freeze_panes = "A2"

            if ws.max_row > 1 and ws.max_column > 0:
                ws.auto_filter.ref = ws.dimensions

            ws.row_dimensions[1].height = 28

            col_headers = {}
            for col_idx in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=1, column=col_idx).value
                if cell_val is not None:
                    col_headers[col_idx] = str(cell_val).strip()

            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = header_border

            for row_idx in range(2, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 20
                is_even = (row_idx % 2 == 0)
                row_fill = white_fill if is_even else zebra_fill

                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = data_font
                    cell.fill = row_fill
                    cell.border = thin_border

                    col_name = col_headers.get(col_idx, "")

                    if col_name in currency_cols:
                        cell.number_format = "#,##0 €"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_name in diff_currency_cols:
                        cell.number_format = "+#,##0 €;-#,##0 €;0 €"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_name in pct_cols:
                        cell.number_format = "0.0%"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_name in decimal_cols:
                        cell.number_format = "0.00"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_name in integer_cols:
                        cell.number_format = "#,##0"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_name in center_cols:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")

            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for row_idx in range(1, min(ws.max_row + 1, 80)):
                    v = ws.cell(row=row_idx, column=col_idx).value
                    if v is not None:
                        max_len = max(max_len, len(str(v)))
                col_header = col_headers.get(col_idx, "")
                header_len = len(col_header)
                target_width = max(max_len, header_len) + 4
                ws.column_dimensions[col_letter].width = max(10, min(target_width, 50))

        wb.save(excel_path)
    except Exception as e:
        print(f"⚠️ Warning applying Excel formatting: {e}")

def orchestrate_pipeline(extract: bool = True) -> pd.DataFrame:
    """
    Orchestrates the data extraction and transformation pipeline.
    
    Args:
        extract (bool): If True, runs the data extraction process. 
                        If False, skips extraction and loads data from existing CSVs.
    
    Returns:
        pd.DataFrame: The final master dataframe ready for analysis.
    """
    if extract:
        extract_and_save_data()
    
    # 1. Load Data
    data = import_data()
    
    # 2. Rename and Normalize Data
    data = rename_and_normalize_columns(data)
    
    # 3. Process Auxiliary Data
    data = process_comuniate(data)
    data = process_odds(data)
    
    # 4. Consolidate Player Data
    df_players_total = consolidate_player_data(data)
    
    # 5. Feature Engineering
    if df_players_total is not None and not df_players_total.empty:
        df_master = feature_engineering(df_players_total)
        print_step("T6", "Saving Master DataFrames (players_transformed.csv & _master.xlsx)")
        df_master.to_csv('./data/players_transformed.csv', index=False)

        # Multi-sheet Excel workbook (_master.xlsx)
        with pd.ExcelWriter('./data/_master.xlsx', engine='openpyxl') as writer:
            # Sheet 1: user_info
            df_user = data.get('user_info', pd.DataFrame())
            if df_user.empty and os.path.exists('./data/user_info.csv'):
                df_user = pd.read_csv('./data/user_info.csv')
            if not df_user.empty:
                df_user.to_excel(writer, sheet_name="user_info", index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name="user_info", index=False)

            # Sheet 2: active_events (Rounds schedule)
            df_events = data.get('active_events', pd.DataFrame())
            if df_events.empty and os.path.exists('./data/active_events.csv'):
                df_events = pd.read_csv('./data/active_events.csv')
            if not df_events.empty:
                df_events.to_excel(writer, sheet_name="active_events", index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name="active_events", index=False)

            # Sheet 3: players_transformed
            df_master.to_excel(writer, sheet_name="players_transformed", index=False)

            # Sheet 4: rival_financials
            df_rival = data.get('rival_financials', pd.DataFrame())
            if df_rival.empty and os.path.exists('./data/rival_financials.csv'):
                df_rival = pd.read_csv('./data/rival_financials.csv')
            if not df_rival.empty:
                df_rival.to_excel(writer, sheet_name="rival_financials", index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name="rival_financials", index=False)

            # Sheet 5: news
            df_news = data.get('news', pd.DataFrame())
            if df_news.empty and os.path.exists('./data/news.csv'):
                df_news = pd.read_csv('./data/news.csv')
            if not df_news.empty:
                df_news.to_excel(writer, sheet_name="news", index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name="news", index=False)

        # Apply executive formatting
        format_master_excel('./data/_master.xlsx')

        return df_master
    
    return df_players_total
